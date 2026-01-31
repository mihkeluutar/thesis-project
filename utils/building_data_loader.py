"""
Helper module to load campus building data into a unified dataframe.

Uses the same logic as data-exploration.ipynb: reads the overview and building
codes Excel files, then for a given building code loads all measurement CSVs
from data/campus-data/<building>/<year>/ and merges them into one hourly
time-series dataframe. Optionally merges Tallinn weather data (keskkonnaportaal,
~5 km from campus) so each building dataframe includes weather columns.

Usage:
    from building_data_loader import get_building_dataframe, get_building_overview_info, load_campus_metadata, load_weather_dataframe

    # One-liner: building + weather (default)
    df_u06 = get_building_dataframe("U06")

    # Another building (same structure, same weather)
    df_u01 = get_building_dataframe("U01")

    # Reuse loaded metadata for multiple buildings
    overview_df, building_codes_df = load_campus_metadata()
    weather_df = load_weather_dataframe()  # load once
    df_u06 = get_building_dataframe("U06", overview_df=overview_df, building_codes_df=building_codes_df, weather_df=weather_df)
    df_u01 = get_building_dataframe("U01", overview_df=overview_df, building_codes_df=building_codes_df, weather_df=weather_df)

    # Building only, no weather
    df = get_building_dataframe("U06", include_weather=False)
"""

from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# Default paths (relative to project root or cwd)
DEFAULT_CAMPUS_DATA_DIR = Path("data/campus-data")
CAMPUS_OVERVIEW_FILENAME = "andmed ulevaade.xlsx"
CAMPUS_BUILDING_CODES_FILENAME = "hooned koodid.xlsx"
EHR_COLUMN_LETTER = "B"

# Weather: Tallinn (keskkonnaportaal). Measurement point ~5 km from campus, suitable for all buildings.
DEFAULT_WEATHER_DATA_PATH = Path("data/keskkonnaportaal/tallinn-harku_f_kliima_tund.csv")

# Measurement point detail types (suffixes in CSV filenames)
MEASUREMENT_POINT_DETAILS = ["näit", "pealevoolu temp", "tagasivoolu temp"]
DEFAULT_YEARS = ["2022", "2023", "2024"]


def load_campus_metadata(
    campus_data_dir: Optional[Path] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load the overview and building codes Excel files, including EHR hyperlinks.

    Returns:
        (overview_df, building_codes_df). building_codes_df includes an
        "EHR register URL" column with hyperlinks extracted from the Excel.
    """
    campus_data_dir = campus_data_dir or DEFAULT_CAMPUS_DATA_DIR
    overview_path = campus_data_dir / CAMPUS_OVERVIEW_FILENAME
    building_codes_path = campus_data_dir / CAMPUS_BUILDING_CODES_FILENAME

    overview_df = pd.read_excel(overview_path)
    building_codes_df = pd.read_excel(building_codes_path)

    # Extract hyperlinks from Excel (column B = "EHR register")
    wb = load_workbook(building_codes_path, data_only=False)
    ws = wb.active
    hyperlinks = []
    for row_idx in range(2, ws.max_row + 1):
        cell = ws[f"{EHR_COLUMN_LETTER}{row_idx}"]
        url = cell.hyperlink.target if cell.hyperlink else None
        hyperlinks.append(url)
    building_codes_df["EHR register URL"] = hyperlinks[: len(building_codes_df)]

    return overview_df, building_codes_df


def load_weather_dataframe(
    weather_data_path: Optional[Path] = None,
) -> pd.DataFrame:
    """
    Load the Tallinn weather data from keskkonnaportaal (semicolon-delimited CSV).

    Same logic as data-exploration.ipynb: build Time from year/month/day/hour,
    pivot by feature (element code), rename columns to English element names.
    Weather station is ~5 km from campus, suitable for all university buildings.

    Returns:
        DataFrame with columns: Time, Air pressure at sea level, Precipitation (hourly sum),
        Relative humidity, Air temperature, etc.
    """
    path = weather_data_path or DEFAULT_WEATHER_DATA_PATH
    weather_data = pd.read_csv(path, delimiter=";")

    weather_data["Time"] = pd.to_datetime(
        weather_data["aasta - Year of measurement, UTC time"].astype(str)
        + "-"
        + weather_data["kuu - Month of measurement, UTC time"].astype(str).str.zfill(2)
        + "-"
        + weather_data["paev - Day of measurement, UTC time"].astype(str).str.zfill(2)
        + " "
        + weather_data["tund - Hour of measurement, UTC time"].astype(str).str.zfill(2)
        + ":00:00"
    )
    weather_data["value"] = (
        weather_data["vaartus - The measured value"]
        .astype(str)
        .str.replace(",", ".")
        .astype(float)
    )
    weather_data["feature"] = weather_data["element_kood - Element code, local"]

    weather_df = weather_data.pivot_table(
        index="Time",
        columns="feature",
        values="value",
        aggfunc="mean",
    ).reset_index()

    feature_mapping = weather_data[
        [
            "element_kood - Element code, local",
            "element_nimi_eng - Element name (eng)",
        ]
    ].drop_duplicates()
    feature_mapping = dict(
        zip(
            feature_mapping["element_kood - Element code, local"],
            feature_mapping["element_nimi_eng - Element name (eng)"],
        )
    )
    weather_df.columns = ["Time"] + [
        feature_mapping.get(col, col) for col in weather_df.columns[1:]
    ]
    return weather_df


def get_building_overview_info(
    building_code: str,
    overview_df: pd.DataFrame,
    building_codes_df: pd.DataFrame,
) -> dict:
    """
    Get metadata and year-coverage summary for a building (same logic as notebook).

    Returns a dict with: building_name, building_kood, ehr_url, year_data_ratios,
    measurement_points (list of "ID convert" values with point_id).
    """
    building_data = building_codes_df[
        building_codes_df["kood"].str.contains(building_code, na=False)
    ]
    if building_data.empty:
        raise ValueError(f"Building code '{building_code}' not found in building codes.")
    building_data_row = building_data.iloc[0]

    building_overview_data = overview_df[overview_df["Hoone"] == building_code]
    if building_overview_data.empty:
        raise ValueError(
            f"No overview rows for building '{building_code}'. Check 'Hoone' column."
        )

    year_columns = [2022, 2023, 2024]
    year_data_ratios = {}
    for year in year_columns:
        non_null = building_overview_data[year].notna().sum()
        total = len(building_overview_data)
        year_data_ratios[year] = f"{non_null}/{total}"

    measurement_points = []
    for point_name in building_overview_data["ID convert"].unique():
        point_id = building_overview_data[
            building_overview_data["ID convert"] == point_name
        ]["point_id"].values[0]
        measurement_points.append({"id_convert": point_name, "point_id": point_id})

    return {
        "building_name": building_data_row["Hoone"],
        "building_kood": building_data_row["kood"],
        "ehr_url": building_data_row.get("EHR register URL"),
        "year_data_ratios": year_data_ratios,
        "measurement_points": measurement_points,
    }


def get_building_dataframe(
    building_code: str,
    campus_data_dir: Optional[Path] = None,
    overview_df: Optional[pd.DataFrame] = None,
    building_codes_df: Optional[pd.DataFrame] = None,
    years: Optional[list] = None,
    resample_hourly: bool = True,
    add_näit_delta: bool = True,
    include_weather: bool = True,
    weather_data_path: Optional[Path] = None,
    weather_df: Optional[pd.DataFrame] = None,
    verbose: bool = False,
) -> pd.DataFrame:
    """
    Build a single dataframe for one university building with all measurement series.

    Optionally merges Tallinn weather data (keskkonnaportaal, ~5 km from campus).

    Logic (from data-exploration.ipynb):
      - Filter overview by Hoone == building_code and get "ID convert" measurement points.
      - For each (year, measurement_point, measurement_point_detail), load CSV from
        campus_data_dir / building_code / year / {building_code}_{point}_{detail}_{year}.csv
      - Combine Time + Value into columns, merge years per column (groupby Time, mean),
        then merge all columns into one dataframe.
      - Optionally resample to 1h (mean) and add näit_delta columns (diff of "näit").

    Parameters
    ----------
    building_code : str
        Short building code, e.g. "U06". Must match "Hoone" in overview and folder name.
    campus_data_dir : Path, optional
        Base path for campus data (contains andmed ulevaade.xlsx, hooned koodid.xlsx,
        and subdirs like U06/2022/, U06/2023/, ...). Default: data/campus-data.
    overview_df, building_codes_df : DataFrame, optional
        If provided, use these instead of loading from campus_data_dir. Useful to load
        metadata once and call get_building_dataframe for multiple buildings.
    years : list of str, optional
        Years to load, e.g. ["2022", "2023", "2024"]. Default: ["2022", "2023", "2024"].
    resample_hourly : bool
        If True, resample to 1h resolution (mean). Default True.
    add_näit_delta : bool
        If True, add columns like BHB01_näit_delta = diff(BHB01_näit). Default True.
    include_weather : bool
        If True, merge Tallinn weather data (left join on Time). Default True.
    weather_data_path : Path, optional
        Path to keskkonnaportaal CSV. Used only if include_weather=True and
        weather_df is None. Default: data/keskkonnaportaal/tallinn-harku_f_kliima_tund.csv.
    weather_df : DataFrame, optional
        Pre-loaded weather dataframe. If provided, used instead of loading from path.
    verbose : bool
        If True, print progress (files loaded, row counts). Default False.

    Returns
    -------
    pd.DataFrame
        Columns: Time, building measurement columns (BHB01_näit, ...), optionally
        näit_delta columns, and if include_weather=True: Air temperature,
        Relative humidity, Precipitation (hourly sum), etc.
    """
    campus_data_dir = campus_data_dir or DEFAULT_CAMPUS_DATA_DIR
    years = years or DEFAULT_YEARS

    if overview_df is None or building_codes_df is None:
        overview_df, building_codes_df = load_campus_metadata(campus_data_dir)

    building_overview_data = overview_df[overview_df["Hoone"] == building_code]
    if building_overview_data.empty:
        raise ValueError(
            f"No overview rows for building '{building_code}'. "
            "Check that 'Hoone' in andmed ulevaade.xlsx matches the building code."
        )

    measurement_points_full = building_overview_data["ID convert"].unique()
    measurement_points = [
        point.replace(f"{building_code}.", "") for point in measurement_points_full
    ]

    column_data = {}
    total_files = len(years) * len(measurement_points) * len(MEASUREMENT_POINT_DETAILS)
    if verbose:
        print(f"Checking up to {total_files} files...")

    file_count = 0
    for year in years:
        for measurement_point in measurement_points:
            for measurement_point_detail in MEASUREMENT_POINT_DETAILS:
                file_path = (
                    campus_data_dir
                    / building_code
                    / year
                    / f"{building_code}_{measurement_point}_{measurement_point_detail}_{year}.csv"
                )
                if not file_path.exists():
                    continue
                file_count += 1
                col_name = f"{measurement_point}_{measurement_point_detail}"

                df = pd.read_csv(file_path)
                df = df[["Time", "Value"]].copy()
                df["Time"] = pd.to_datetime(df["Time"])
                df = df.rename(columns={"Value": col_name})

                if col_name not in column_data:
                    column_data[col_name] = []
                column_data[col_name].append(df)
                if verbose:
                    print(f"  [{file_count}] {col_name} ({year}): {len(df)} rows")

    if verbose:
        print(f"Loaded {file_count} files into {len(column_data)} columns.")

    # Combine years per column: concat then groupby Time, mean
    column_series = {}
    for col_name, dfs in column_data.items():
        combined = pd.concat(dfs, ignore_index=True)
        series = combined.groupby("Time")[col_name].mean()
        column_series[col_name] = series
        if verbose:
            print(f"  {col_name}: {len(series)} unique timestamps")

    dataframe = pd.DataFrame(column_series)
    dataframe = dataframe.sort_index().reset_index()

    if resample_hourly:
        df_resampled = dataframe.set_index("Time")
        df_resampled = df_resampled.resample("1h").mean()
        df_resampled = df_resampled.reset_index()
        dataframe = df_resampled
        if verbose:
            print(f"Resampled to 1h: {len(dataframe)} rows")

    if add_näit_delta:
        nait_cols = [
            col
            for col in dataframe.columns
            if col.endswith("_näit") and "delta" not in col
        ]
        for col in nait_cols:
            new_col = col.replace("näit", "näit_delta")
            dataframe[new_col] = dataframe[col].diff()
        if verbose and nait_cols:
            print(f"Added {len(nait_cols)} näit_delta columns.")

    if include_weather:
        if weather_df is None:
            weather_df = load_weather_dataframe(weather_data_path)
            if verbose:
                print(f"Weather: loaded from {weather_data_path or DEFAULT_WEATHER_DATA_PATH}, {len(weather_df)} rows.")
        dataframe = pd.merge(dataframe, weather_df, on="Time", how="left")
        if verbose:
            print(f"Merged weather: {dataframe.shape[1]} columns total.")

    return dataframe
